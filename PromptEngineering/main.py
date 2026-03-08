from fastapi import FastAPI
from fastapi.responses import RedirectResponse
from pydantic import BaseModel
from agent_calling import groq_chat_completion
import gradio as gr
import asyncio
from openpyxl import Workbook, load_workbook
import os

app = FastAPI()

# Excel logging function
def log_to_excel(question: str, answer: str, filename: str = "qa_log.xlsx"):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "QA Log"
        ws.append(["Question", "Answer"])
        wb.save(filename)
    wb = load_workbook(filename)
    ws = wb.active
    ws.append([question, answer])
    wb.save(filename)


class MessageRequest(BaseModel):
    message: str


@app.get("/")
async def root():
    return RedirectResponse(url="/gradio")


@app.post("/cli-command")
async def message(request: MessageRequest):
    response = await groq_chat_completion(request.message)
    if isinstance(response, dict) and response.get("error"):
        return {
            "user_message": request.message,
            "grok_response": "",
            "error": response.get("error")
        }

    return {
        "user_message": request.message,
        "grok_response": response.get("choices", [{}])[0].get("message", {}).get("content", "")
    }



# Gradio Interface
def generate_cli_command(user_message: str) -> str:
    """Convert natural language to CLI command using Grok AI."""
    try:
        response = asyncio.run(groq_chat_completion(user_message))
        if isinstance(response, dict) and response.get("error"):
            return f"Error: {response.get('error')}"

        cli_command = response.get("choices", [{}])[0].get("message", {}).get("content", "")
        log_to_excel(user_message, cli_command)
        return cli_command
    except Exception as e:
        return f"Error: {str(e)}"


# Create Gradio interface
gradio_interface = gr.Interface(
    fn=generate_cli_command,
    inputs=gr.Textbox(
        label="Describe what you want to do",
        placeholder="e.g., 'list all files in the current directory'",
        lines=3
    ),
    outputs=gr.Textbox(
        label="Generated CLI Command",
        interactive=False,
        lines=3
    ),
    title="CLI Command Generator with Grok AI",
    description="Convert natural language descriptions into executable terminal commands",
    examples=[
        ["list all files in the current folder"],
        ["create a new folder called reports"],
        ["find all Python files recursively"],
        ["show processes using port 8080"],
    ]
)

# Mount Gradio app
app = gr.mount_gradio_app(app, gradio_interface, path="/gradio")

print("DEBUG: main.py loaded")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8080)